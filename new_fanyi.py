import pandas as pd
import openpyxl
from googletrans import Translator
import requests
import json
from volcengine.ApiInfo import ApiInfo
from volcengine.Credentials import Credentials
from volcengine.ServiceInfo import ServiceInfo
from volcengine.base.Service import Service
import configparser
from translation_service_provider import *


# 计算文件的字符总数
def count_characters(df):
    total_characters = 0
    for column in df.columns:
        for _, row in df.iterrows():
            if isinstance(row[column], str):
                total_characters += len(row[column])
    return total_characters


def translate_text(text, service_instance, service_provider):
    if text is None or pd.isna(text):
        return ""

    try:
        # 火山翻译服务
        if isinstance(service_instance, Service):
            body = {
                'TargetLanguage': 'en',
                'TextList': [text]
            }
            res = service_instance.json('translate', {}, json.dumps(body))
            result = json.loads(res)
            if 'TranslationList' in result and result['TranslationList']:
                return result['TranslationList'][0]['Translation']
            else:
                return text

        # 腾讯翻译服务
        elif isinstance(service_instance, TC3Client):  # 确保在这里判断的是 service_instance
            # print(f"service_instance的类型是: {type(service_instance)}")
            print(f"即将调用construct_request方法，文本: {text}")

            res = service_instance.construct_request([text])  # 在 service_instance 上调用 construct_request
            print(f"construct_request调用成功，返回结果: {res}")

            result = json.loads(res)
            # print(f"翻译服务返回结果：{result}")
            if 'Response' in result and 'TargetTextList' in result['Response']:
                return result['Response']['TargetTextList'][0]
            else:
                print("翻译结果格式不正确，返回原始文本")
                return text

    except Exception as e:
        import traceback
        print(f"翻译文本时出现错误：{e}")
        print(traceback.format_exc())
        return text  # 返回原始文本以防止异常导致程序中断


def process_excel(config_name):
    global translation_result
    config = configparser.ConfigParser()
    config.read('config.ini')
    service_provider = TranslationServiceProvider()
    service_instance = service_provider.get_service_info(config_name)


    input_file = service_provider.config.get(config_name, 'input_file_path')
    output_file = service_provider.config.get(config_name, 'output_file_path')

    workbook = openpyxl.load_workbook(input_file)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.data_type = 'str'

    # service_instance = get_translation_service(config_name)
    worksheet = workbook.active
    merged_ranges = worksheet.merged_cells.ranges
    print("读取到的合并单元格信息：")
    for merged_range in merged_ranges:
        print(
            f"起始行：{merged_range.min_row}，结束行：{merged_range.max_row}，起始列：{merged_range.min_col}，结束列：{merged_range.max_col}")

    df_original = pd.read_excel(input_file, engine='openpyxl', keep_default_na=False)
    columns_with_content = [column for column in df_original.columns if
                            any(isinstance(row[column], str) for _, row in df_original.iterrows())]
    total_characters = count_characters(df_original[columns_with_content])
    print(f"总字符数为：{total_characters}")
    df_translated = df_original.copy()
    total_cells = len(df_original) * len(columns_with_content)
    translated_cells = 0
    translated_characters = 0
    for column in columns_with_content:
        last_value = None
        for index, row in df_translated.iterrows():
            cell_value = row[column]
            if not isinstance(cell_value, str):
                cell_value = str(cell_value)

            if pd.isna(cell_value):
                df_translated.at[index, column] = last_value
            else:
                translation_result = translate_text(cell_value, service_instance, service_provider)
                df_translated.at[index, column] = translation_result
                is_merged_cell = False
                for merged_range in merged_ranges:
                    if index >= merged_range.min_row and index <= merged_range.max_row and column == merged_range.min_col:
                        merged_cell_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
                        if not isinstance(merged_cell_value, str):
                            merged_cell_value = str(merged_cell_value)
                        translation_result = translate_text(merged_cell_value, service_instance, service_provider)
                        for r in range(merged_range.min_row, merged_range.max_row + 1):
                            for c in range(merged_range.min_col, merged_range.max_col + 1):
                                df_translated.at[r - 1, c - 1] = translation_result
                        is_merged_cell = True
                        break
                if not is_merged_cell:
                    translation_result = translate_text(cell_value, service_instance, service_provider)
                    df_translated.at[index, column] = translation_result
                if translation_result is None:
                    translation_result = ""
                translated_characters += len(translation_result)
                # last_value = translation_result
                translated_cells += 1
                # translated_characters += len(translation_result)
                print(
                    f"翻译进度：{translated_cells}/{total_cells}（{translated_cells / total_cells * 100:.2f}%），已翻译字符数：{translated_characters}/{total_characters}")
    print("翻译后的数据框内容：", df_translated)
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df_translated.to_excel(writer, index=False)
    new_workbook = writer.book
    new_worksheet = new_workbook.active
    for merged_range in merged_ranges:
        new_worksheet.merge_cells(start_row=merged_range.min_row, start_column=merged_range.min_col,
                                  end_row=merged_range.max_row, end_column=merged_range.max_col)
    writer.close()


process_excel('hs_api')
